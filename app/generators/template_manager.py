# template_manager.py - Loading, cloning, and formatting the Excel templates
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.cell_range import CellRange
from copy import copy
import shutil
import os

def copy_cell_style(src_cell, dest_cell):
    """
    Copies style (font, alignment, border, fill, number_format) from src_cell to dest_cell.
    """
    if src_cell.has_style:
        dest_cell.font = copy(src_cell.font)
        dest_cell.alignment = copy(src_cell.alignment)
        dest_cell.border = copy(src_cell.border)
        dest_cell.fill = copy(src_cell.fill)
        dest_cell.number_format = copy(src_cell.number_format)

def adjust_rows_in_sheet(ws, required_rows, first_data_row=8, total_row_label_cell="A21"):
    """
    Adjusts the number of rows in the table.
    - If required_rows > available placeholder rows (20 - 8 + 1 = 13), inserts rows before the G.TOTAL row.
    - If required_rows < 13, deletes unused placeholder rows so that the layout is clean.
    - Copies styling from the first_data_row to all new rows.
    - Manually shifts merged cell ranges to prevent openpyxl corrupting them.
    """
    placeholder_rows = 13 # rows 8 to 20
    diff = required_rows - placeholder_rows
    
    # Identify the current total row index
    total_row_idx = 21
    for r in range(first_data_row, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "G.TOTAL":
            total_row_idx = r
            break
            
    if diff > 0:
        # Shift merged cell ranges manually before inserting rows
        merged_ranges = list(ws.merged_cells.ranges)
        for r_range in merged_ranges:
            if r_range.min_row >= total_row_idx:
                ws.merged_cells.remove(r_range)
                r_range.shift(row_shift=diff)
                ws.merged_cells.add(r_range)
                
        # Insert rows right before the total row
        ws.insert_rows(total_row_idx, diff)
        # Copy styles from the first data row (row 8) to the newly inserted rows
        for r in range(total_row_idx, total_row_idx + diff):
            for c in range(1, ws.max_column + 1):
                src_cell = ws.cell(row=first_data_row, column=c)
                dest_cell = ws.cell(row=r, column=c)
                copy_cell_style(src_cell, dest_cell)
    elif diff < 0:
        # Delete unused rows right before the total row
        rows_to_delete = -diff
        start_delete_row = total_row_idx - rows_to_delete
        
        # Shift merged ranges UP (diff is negative)
        merged_ranges = list(ws.merged_cells.ranges)
        for r_range in merged_ranges:
            if r_range.min_row >= total_row_idx:
                ws.merged_cells.remove(r_range)
                r_range.shift(row_shift=diff)
                ws.merged_cells.add(r_range)
                
        ws.delete_rows(start_delete_row, rows_to_delete)
        
    # Return the new total row index
    for r in range(first_data_row, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "G.TOTAL":
            return r
    return 21

def format_total_row_formulas(ws, total_row_idx, start_row=8, cols_range=None):
    """
    Adds SUM formulas to the total row cells for the specified columns.
    E.g. =SUM(C8:C20)
    """
    if not cols_range:
        return
        
    for col_let in cols_range:
        cell = ws[f"{col_let}{total_row_idx}"]
        cell.value = f"=SUM({col_let}{start_row}:{col_let}{total_row_idx-1})"
        
        # Ensure it is bold
        cell.font = Font(name=cell.font.name, size=cell.font.size, bold=True)
