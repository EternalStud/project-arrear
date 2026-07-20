from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from app.generators.template_manager import adjust_rows_in_sheet, format_total_row_formulas

def write_salary_arrear_sheet(ws, arrear_result):
    """
    Populates the 'Salary Arrear Format' sheet in the workbook.
    """
    employee = arrear_result["employee"]
    arrear_months = arrear_result["arrear_months"]
    totals = arrear_result["totals"]
    in_words = arrear_result["in_words"]
    
    # 1. Fill Employee Metadata headers
    ws["A3"] = f"NAME OF SCHOOL- {employee.get('school_name') or ''}"
    ws["N3"] = f"BLOCK NAME - {employee.get('block_name') or ''}"
    
    ws["A4"] = f"NAME OF TEACHER- {employee.get('name') or ''}"
    ws["I4"] = f"DESIGNATION- {employee.get('designation') or ''}"
    ws["Q4"] = f"DATE OF JOINING- {employee.get('doj') or ''}"
    
    ws["A5"] = f"PRAN- {employee.get('pran') or ''}"
    ws["I5"] = f"ACCOUN NO.- {employee.get('bank_account') or ''}"
    ws["Q5"] = f"IFSC - {employee.get('ifsc') or ''}"
    
    # 2. Adjust rows dynamically based on the number of arrear months
    num_months = len(arrear_months)
    start_row = 8
    total_row_idx = adjust_rows_in_sheet(ws, num_months, first_data_row=start_row)
    
    # 3. Write data rows
    for idx, item in enumerate(arrear_months):
        r = start_row + idx
        adm = item["admissible"]
        drn = item["drawn"]
        diff = item["difference"]
        
        # Column A: Month
        ws.cell(row=r, column=1, value=item["month_label"])
        # Column B: No of Days
        ws.cell(row=r, column=2, value=adm["days"])
        
        # Columns C-J: Admissible
        ws.cell(row=r, column=3, value=adm["basic"])
        ws.cell(row=r, column=4, value=adm["da"])
        ws.cell(row=r, column=5, value=adm["hra"])
        ws.cell(row=r, column=6, value=adm["ma"])
        ws.cell(row=r, column=7, value=f"=SUM(C{r}:F{r})") # GROSS = Basic + DA + HRA + MA
        ws.cell(row=r, column=8, value=adm["nps"])
        ws.cell(row=r, column=9, value=adm["gis"])
        
        # Net pay formula: gross - nps - gis.
        # But wait! For September, we subtracted PT. So:
        # If PT is 0, Net is just Gross - NPS - GIS.
        # If PT is > 0, Net is Gross - NPS - GIS - PT.
        # To represent this in Excel formula:
        if adm["professional_tax"] > 0:
            ws.cell(row=r, column=10, value=f"=G{r}-H{r}-I{r}-{adm['professional_tax']}")
        else:
            ws.cell(row=r, column=10, value=f"=G{r}-H{r}-I{r}")
            
        # Columns K-R: Drawn
        ws.cell(row=r, column=11, value=drn["basic"])
        ws.cell(row=r, column=12, value=drn["da"])
        ws.cell(row=r, column=13, value=drn["hra"])
        ws.cell(row=r, column=14, value=drn["ma"])
        ws.cell(row=r, column=15, value=f"=SUM(K{r}:N{r})")
        ws.cell(row=r, column=16, value=drn["nps"])
        ws.cell(row=r, column=17, value=drn["gis"])
        
        if drn["professional_tax"] > 0:
            ws.cell(row=r, column=18, value=f"=O{r}-P{r}-Q{r}-{drn['professional_tax']}")
        else:
            ws.cell(row=r, column=18, value=f"=O{r}-P{r}-Q{r}")
            
        # Column S: Paid Arrear (Column 19)
        arrear_drawn = drn.get("arrear_drawn", 0)
        ws.cell(row=r, column=19, value=arrear_drawn)
        
        # Columns T-AA: Difference (using formulas)
        ws.cell(row=r, column=20, value=f"=C{r}-K{r}") # Basic diff
        ws.cell(row=r, column=21, value=f"=D{r}-L{r}") # DA diff
        ws.cell(row=r, column=22, value=f"=E{r}-M{r}") # HRA diff
        ws.cell(row=r, column=23, value=f"=F{r}-N{r}") # MA diff
        ws.cell(row=r, column=24, value=f"=G{r}-O{r}") # Gross diff
        ws.cell(row=r, column=25, value=f"=H{r}-P{r}") # NPS diff
        ws.cell(row=r, column=26, value=f"=I{r}-Q{r}") # GIS diff
        ws.cell(row=r, column=27, value=f"=J{r}-R{r}-S{r}") # Net diff = Admissible Net - Drawn Net - Paid Arrear
        
    # 4. Write SUM formulas to the G.TOTAL row
    sum_cols = [get_column_letter(c) for c in range(3, 28)] # Columns C to AA
    format_total_row_formulas(ws, total_row_idx, start_row=start_row, cols_range=sum_cols)
    
    # 5. Write Net total in words in cell C22 (shifted to total_row_idx + 1)
    in_words_row = total_row_idx + 1
    ws[f"C{in_words_row}"] = f"IN WORDS: {in_words}"
    ws.row_dimensions[in_words_row].height = 22.0
    ws[f"C{in_words_row}"].font = Font(name="Arial", size=10, bold=True)
    
    # 6. Format Signature & Seal row height & alignment (shifted to total_row_idx + 3)
    sig_row = total_row_idx + 3
    ws.row_dimensions[sig_row].height = 45.0
    
    cell_ts = ws[f"A{sig_row}"]
    cell_ts.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell_ts.font = Font(name="Arial", size=10, bold=True)
    
    cell_hm = ws[f"S{sig_row}"]
    cell_hm.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell_hm.font = Font(name="Arial", size=10, bold=True)
