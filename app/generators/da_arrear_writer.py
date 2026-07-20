from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from app.generators.template_manager import adjust_rows_in_sheet, format_total_row_formulas

def write_da_arrear_sheet(ws, arrear_result):
    """
    Populates the 'DA Arrear Format' sheet in the workbook.
    """
    employee = arrear_result["employee"]
    arrear_months = arrear_result["arrear_months"]
    
    # 1. Fill Employee Metadata headers
    ws["A3"] = f"NAME OF SCHOOL- {employee.get('school_name') or ''}"
    ws["I3"] = f"BLOCK NAME - {employee.get('block_name') or ''}"
    
    ws["A4"] = f"NAME OF TEACHER- {employee.get('name') or ''}"
    ws["F4"] = f"DESIGNATION- {employee.get('designation') or ''}"
    ws["K4"] = f"DATE OF JOINING- {employee.get('doj') or ''}"
    
    ws["A5"] = f"PRAN- {employee.get('pran') or ''}"
    ws["F5"] = f"ACCOUN NO.- {employee.get('bank_account') or ''}"
    ws["K5"] = f"IFSC - {employee.get('ifsc') or ''}"
    
    # 2. Adjust rows dynamically based on the number of arrear months
    num_months = len(arrear_months)
    start_row = 8
    total_row_idx = adjust_rows_in_sheet(ws, num_months, first_data_row=start_row)
    
    # 3. Write data rows
    for idx, item in enumerate(arrear_months):
        r = start_row + idx
        adm = item["admissible"]
        drn = item["drawn"]
        
        # Column A: Month
        ws.cell(row=r, column=1, value=item["month_label"])
        # Column B: No of Days
        ws.cell(row=r, column=2, value=adm["days"])
        
        # Columns C-G: Admissible
        ws.cell(row=r, column=3, value=adm["basic"])
        ws.cell(row=r, column=4, value=adm["da"])
        ws.cell(row=r, column=5, value=f"=C{r}+D{r}") # GROSS AMOUNT = Basic + DA
        ws.cell(row=r, column=6, value=adm["nps"])
        ws.cell(row=r, column=7, value=f"=E{r}-F{r}") # NET AMOUNT = Gross - NPS
        
        # Columns H-L: Drawn
        ws.cell(row=r, column=8, value=drn["basic"])
        ws.cell(row=r, column=9, value=drn["da"])
        ws.cell(row=r, column=10, value=f"=H{r}+I{r}")
        ws.cell(row=r, column=11, value=drn["nps"])
        ws.cell(row=r, column=12, value=f"=J{r}-K{r}")
        
        # Columns M-Q: Difference (using formulas)
        ws.cell(row=r, column=13, value=f"=C{r}-H{r}") # Basic diff
        ws.cell(row=r, column=14, value=f"=D{r}-I{r}") # DA diff
        ws.cell(row=r, column=15, value=f"=E{r}-J{r}") # Gross diff
        ws.cell(row=r, column=16, value=f"=F{r}-K{r}") # NPS diff
        ws.cell(row=r, column=17, value=f"=G{r}-L{r}") # Net diff
        
    # 4. Write SUM formulas to the G.TOTAL row
    sum_cols = [get_column_letter(c) for c in range(3, 18)] # Columns C to Q
    format_total_row_formulas(ws, total_row_idx, start_row=start_row, cols_range=sum_cols)
    
    # 5. Format Signature & Seal row height & alignment (shifted to total_row_idx + 2)
    sig_row = total_row_idx + 2
    ws.row_dimensions[sig_row].height = 45.0
    
    cell_ts = ws[f"A{sig_row}"]
    cell_ts.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell_ts.font = Font(name="Arial", size=10, bold=True)
    
    cell_hm = ws[f"L{sig_row}"]
    cell_hm.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    cell_hm.font = Font(name="Arial", size=10, bold=True)
