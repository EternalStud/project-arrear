import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

df = pd.read_excel("salary_master.xlsx")

for name, group in df.groupby("Name"):

    pran = str(group["Pran"].iloc[0])
    pan = group["Pan"].iloc[0]
    school = group["School"].iloc[0] if "School" in group.columns else ""

    wb = Workbook()
    ws = wb.active

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    thin = Side(style='thin')
    medium = Side(style='medium')

    border_thin = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ===== TITLE =====
    ws.merge_cells("A2:L2")
    ws["A2"] = "SALARY STATEMENT CUM INCOME TAX CALCULATION(F.Y. - 2025-26 ASSESSMENT YEAR 2026-27)"
    ws["A2"].font = bold
    ws["A2"].alignment = center

    # ===== HEADER =====
    ws["A3"] = "Name of the Employee"
    ws["B3"] = name

    ws.merge_cells("C3:G3")
    ws["C3"] = school
    ws["C3"].alignment = center

    ws["H3"] = "PAN NO."
    ws["I3"] = pan

    # ===== TABLE HEADERS =====
    headers = ["Month","Pay","D.A.","H.R.A.","M.A.","Total","N.P.S","G.L.I","I.T","PROF.TAX","TOTAL","NET PAY"]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = h
        cell.font = bold
        cell.alignment = center
        cell.border = border_thin

    months_order = ["Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec","Jan","Feb"]

    row_start = 5
    totals = {"Basic":0,"DA":0,"HRA":0,"MA":0,"Total":0,"NPS":0,"GLI":0,"PT":0,"Deduction":0,"Net":0}

    # ===== DATA =====
    for i, month in enumerate(months_order):
        r = group[group["Month"] == month].iloc[0]

        values = [
            f"{month}-25" if month not in ["Jan","Feb"] else f"{month}-26",
            r["Basic"], r["DA"], r["HRA"], r["MA"],
            r["Total"], r["NPS"], r["GLI"],
            "",                # I.T always blank
            r["PT"],           # PROF TAX
            r["Deduction"],
            r["Net"]
        ]

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=row_start+i, column=col)
            cell.value = val
            cell.border = border_thin
            cell.alignment = center if col == 1 else right

        totals["Basic"] += r["Basic"]
        totals["DA"] += r["DA"]
        totals["HRA"] += r["HRA"]
        totals["MA"] += r["MA"]
        totals["Total"] += r["Total"]
        totals["NPS"] += r["NPS"]
        totals["GLI"] += r["GLI"]
        totals["PT"] += r["PT"]
        totals["Deduction"] += r["Deduction"]
        totals["Net"] += r["Net"]

    # ===== TOTAL ROW =====
    r = row_start + 12
    ws.cell(row=r, column=1).value = "Total"
    ws.cell(row=r, column=1).font = bold

    total_values = [
        totals["Basic"], totals["DA"], totals["HRA"], totals["MA"],
        totals["Total"], totals["NPS"], totals["GLI"],
        "", totals["PT"], totals["Deduction"], totals["Net"]
    ]

    for col, val in enumerate(total_values, start=2):
        cell = ws.cell(row=r, column=col)
        cell.value = val
        cell.font = bold
        cell.border = border_thin
        cell.alignment = right

    # ===== INCOME TAX SECTION =====
    start = r + 2

    ws.merge_cells(start_row=start, start_column=1, end_row=start, end_column=12)
    ws.cell(row=start, column=1).value = "INCOME TAX CALCULATION SHEET F.Y.(2025-2026)"
    ws.cell(row=start, column=1).alignment = center
    ws.cell(row=start, column=1).font = bold

    income = totals["Total"]
    std_ded = 75000
    taxable = income - std_ded

    tax_5 = max(0, min(taxable-400000, 400000)) * 0.05 if taxable > 400000 else 0

    rows = [
        ("1","INCOME FROM SALARY", income),
        ("2","INCOME IN MULTIPLE OF 10", income),
        ("3","LESS : STANDARD DEDUCTION", std_ded),
        ("4","TAXABLE INCOME", taxable),
        ("5","TAX UP TO 400000","NIL"),
        ("","TAX UP TO NEXT 400001 TO 800000 @ 5%", round(tax_5)),
        ("6","TOTAL TAX", round(tax_5)),
        ("7","LESS REBATE U/S 87A", round(tax_5)),
        ("8","NET INCOME TAX", 0),
        ("9","ADD SURCHARGE @ 4%", 0),
        ("10","TOTAL PAYABLE INCOME TAX", 0),
        ("11","LESS REBATE U/S 89(i)", 0),
        ("12","TAX PAID UP TO JANUARY 2026", 0),
        ("13","LESS TAX DEPOSITED VIDE CHALLAN", 0),
        ("14","TOTAL REFUNDABLE TAX IN 2025-26", 0)
    ]

    for i, (a,b,c) in enumerate(rows):
        ws.cell(row=start+1+i, column=1).value = a
        ws.cell(row=start+1+i, column=2).value = b
        ws.cell(row=start+1+i, column=12).value = c

    # ===== SIGNATURE =====
    ws.cell(row=start+18, column=2).value = "Signature of Employee"
    ws.cell(row=start+18, column=9).value = "DISTRICT PROGRAMME OFFICER\n(ESTABLISHMENT)\nMUZAFFARPUR"

    # WIDTH
    widths = [10,12,12,12,10,14,12,10,8,14,12,12]
    for i,w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=4,column=i).column_letter].width = w

    filename = f"SST_{'_'.join(name.split()).upper()}_{pran}.xlsx"
    wb.save(filename)

    print(f"✅ Created: {filename}")